import openpyxl
import json

# Função para converter os dados de uma linha da planilha para um dicionário
def linha_para_dicionario(sheet, linha):
    dados = {}
    for coluna in range(1, sheet.max_column + 1):
        titulo_coluna = sheet.cell(row=1, column=coluna).value
        valor_celula = sheet.cell(row=linha, column=coluna).value

        # Ajusta para tratar os valores no JSON
        if valor_celula is not None:
            # Para anos ou valores numéricos, manter como está, exceto por espaços extras
            if isinstance(valor_celula, (int, float)):
                valor_celula = str(valor_celula).replace(" ", "")
            elif isinstance(valor_celula, str):
                valor_celula = valor_celula.strip()  # Remove espaços no início e no fim

        dados[titulo_coluna] = valor_celula

    return dados

# Carrega a pasta de trabalho do Excel com fórmulas avaliadas
try:
    workbook = openpyxl.load_workbook('Desempenho dos técnicos da seleção.xlsx', data_only=True)
except FileNotFoundError:
    print("Erro: Arquivo não encontrado. Verifique o nome e o caminho do arquivo.")
    exit()

# Verifica se as abas "Geral" e "Desempenho" existem
abas_necessarias = ["Geral", "Desempenho"]
for aba in abas_necessarias:
    if aba not in workbook.sheetnames:
        print(f"Erro: A aba '{aba}' não foi encontrada no arquivo Excel.")
        exit()

# Cria o objeto dadosSelecao
dadosSelecao = {
    "geral": [],
    "desempenho": [],
}

# Lê os dados da guia "Geral"
sheet_geral = workbook['Geral']
for linha in range(2, sheet_geral.max_row + 1):
    dadosSelecao["geral"].append(linha_para_dicionario(sheet_geral, linha))

# Lê os dados da guia "Desempenho", incluindo as colunas Adversárias e Top5
sheet_desempenho = workbook['Desempenho']
for linha in range(2, sheet_desempenho.max_row + 1):
    dados = linha_para_dicionario(sheet_desempenho, linha)
    
    # Verifica se as colunas Adversárias e Top5 estão presentes
    if "Adversárias" not in dados:
        dados["Adversárias"] = 0  # Valor padrão se a coluna não existir
    if "Top5" not in dados:
        dados["Top5"] = 0  # Valor padrão se a coluna não existir

    dadosSelecao["desempenho"].append(dados)

# Imprime os dados em formato JSON para um arquivo
try:
    with open('dados.json', 'w', encoding='utf-8') as f:
        json.dump(dadosSelecao, f, ensure_ascii=False, indent=4)
    print("Dados exportados para dados.json com sucesso!")
except Exception as e:
    print(f"Erro ao salvar o arquivo JSON: {e}")

